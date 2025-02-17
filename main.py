import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

book = openpyxl.load_workbook('D:\\attendance.xlsx')
sheet = book['Sheet1']
total_rows = sheet.max_row
total_columns = sheet.max_column

staff_emails = ['erakshaya485@gmail.com', 'yyyyyyyy@gmail.com']
subjects = {1: 'CI', 2: 'Python', 3: 'Data Mining'}
warnings = {
    1: "Warning! Only one more absence allowed for CI class.",
    2: "Warning! Only one more absence allowed for Python class.",
    3: "Warning! Only one more absence allowed for Data Mining class."
}

def save_file():
    book.save(r'D:\\attendance.xlsx')
    print("Attendance record updated.")

def send_email(recipients, subject, message):
    sender_email = 'crazygirlaks@gmail.com'
    password = 'ERAkshaya485'
    server = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    server.starttls()
    server.login(sender_email, password)
    
    for recipient in recipients:
        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg.attach(MIMEText(message, 'plain'))
        server.sendmail(sender_email, recipient, msg.as_string())
    
    server.quit()
    print(f"Email sent to {', '.join(recipients)}")

def check_attendance(absent_days, student_rows, subject_code):
    students_to_warn = []
    critical_students = []
    student_rolls = ""
    subject = subjects[subject_code]

    for i, days in enumerate(absent_days):
        if days == 2:
            students_to_warn.append(sheet.cell(row=student_rows[i], column=2).value)
        elif days > 2:
            student_rolls += str(sheet.cell(row=student_rows[i], column=1).value) + ", "
            critical_students.append(sheet.cell(row=student_rows[i], column=2).value)
    
    if students_to_warn:
        send_email(students_to_warn, "Attendance Warning", warnings[subject_code])
    
    if critical_students:
        send_email(critical_students, "Attendance Deficiency", f"You have exceeded the allowed absences in {subject}.")
        send_email([staff_emails[subject_code-1]], "Critical Attendance Report", f"The following students have exceeded absence limits in {subject}: {student_rolls}")

while True:
    print("1 --> CI\n2 --> Python\n3 --> Data Mining")
    subject_choice = int(input("Enter subject code: "))
    num_absentees = int(input("Enter number of absentees: "))
    absent_students = list(map(int, input("Enter roll numbers: ").split())) if num_absentees > 1 else [int(input("Enter roll number: "))]

    student_rows = []
    absent_days = []

    for roll in absent_students:
        for row in range(2, total_rows + 1):
            if sheet.cell(row=row, column=1).value == roll:
                col = subject_choice + 2
                sheet.cell(row=row, column=col).value += 1
                absent_days.append(sheet.cell(row=row, column=col).value)
                student_rows.append(row)
                save_file()
                break
    
    check_attendance(absent_days, student_rows, subject_choice)
    
    if int(input('Check another subject? 1 --> Yes, 0 --> No: ')) == 0:
        break
